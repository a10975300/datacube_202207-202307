from django.contrib import messages
from django.http import HttpResponseRedirect

import xadmin
from .models import ICs
from xadmin.filters import MultiSelectFieldListFilter
from xadmin.layout import Main, Fieldset, Side, Row
from import_export import resources


class ICsAdmin(object):

    list_display = [
        "ic_IC_Function",
        "ic_Vendor",
        "ic_Vendor_PN",
        "ic_Description",
        "ic_Part_Owner",
        "ic_HP_PN",
        "ic_IC_qty",
        #"ic_Full_Feature",
        #"ic_Defeature",
        "ic_Vendor_2nd",
        #"ic_Vendor_PN_2nd",
        #"ic_Vendor_3rd",
        #"ic_Vendor_PN_3rd",
        #"ic_Vendor_4th",
        #"ic_Vendor_PN_4th",
        #"ic_cratedate"
        ]
    list_filter = [
        ("ic_Cycle", MultiSelectFieldListFilter),
        ("ic_Project_name", MultiSelectFieldListFilter),
        ("ic_HW_PM", MultiSelectFieldListFilter),
        ("ic_ODM", MultiSelectFieldListFilter),
        ("ic_Segment", MultiSelectFieldListFilter),
        ("ic_Chipset_Type", MultiSelectFieldListFilter),
        ("ic_Phase", MultiSelectFieldListFilter),
        ("ic_IC_Function", MultiSelectFieldListFilter),
        ("ic_Vendor", MultiSelectFieldListFilter),
        ("ic_Vendor_PN", MultiSelectFieldListFilter),
        ("ic_Description", MultiSelectFieldListFilter),
        ("ic_Part_Owner", MultiSelectFieldListFilter),
        ("ic_HP_PN", MultiSelectFieldListFilter),
        ("ic_IC_qty", MultiSelectFieldListFilter),
        ("ic_Full_Feature", MultiSelectFieldListFilter),
        ("ic_Defeature", MultiSelectFieldListFilter),
        ("ic_Vendor_2nd", MultiSelectFieldListFilter),
        ("ic_Vendor_PN_2nd", MultiSelectFieldListFilter),
        #("ic_Vendor_3rd", MultiSelectFieldListFilter),
        #("ic_Vendor_PN_3rd", MultiSelectFieldListFilter),
        #("ic_Vendor_4th", MultiSelectFieldListFilter),
        #("ic_Vendor_PN_4th", MultiSelectFieldListFilter),
        #("ic_cratedate"),
        ]

    search_fields = [
        "ic_Cycle",
        "ic_Project_name",
        "ic_HW_PM",
        "ic_ODM",
        "ic_Segment",
        "ic_Chipset_Type",
        "ic_Phase",
        "ic_IC_Function",
        "ic_Vendor",
        "ic_Vendor_PN",
        "ic_Description",
        "ic_Part_Owner",
        "ic_HP_PN",
        "ic_IC_qty",
        "ic_Full_Feature",
        "ic_Defeature",
        "ic_Vendor_2nd",
        "ic_Vendor_PN_2nd",
        "ic_Vendor_3rd",
        "ic_Vendor_PN_3rd",
        "ic_Vendor_4th",
        "ic_Vendor_PN_4th",
        "ic_cratedate"
        ]

    list_display_links = ["ic_IC_Function"]
    list_per_page = 50
    date_hierarchy = 'ic_cratedate'
    list_editable = 'ic_IC_Function'
    model_icon = 'fa fa-bandcamp'


    # configuration for import and export data
    class Issue_import_export_Resource(resources.ModelResource):
        class Meta:
            model = ICs
    import_export_args = {'export_resource_class':Issue_import_export_Resource}
    import_excel = True

    def post(self, request, *args, **kwargs):
        if 'excel' in request.FILES:
            ic_report = request.FILES.get('excel')
            from openpyxl import load_workbook
            workbook = load_workbook(ic_report, data_only=True)
            #取得IC EXCEL Sheet name
            sheet_names = workbook.sheetnames
            ic_sheet_name = sheet_names[0]

            # 调用方法处理导入数据 IC_Import_to_Database.excel_to_database
            from .views import IC_Import_to_Database
            IC_Import_to_Database.excel_to_database(self, workbook,ic_sheet_name)

            messages.success(request, "IC report was imported successfully.")
            return HttpResponseRedirect('/scpe/ic/ics/')
        return super(ICsAdmin, self).post(request, *args, **kwargs)


    # 配置 编辑页面
    def get_form_layout(self):
        # if self.org_obj:
        self.form_layout = (
            Main(
                Fieldset('',
                         Row('ic_Cycle', 'ic_Project_name'),
                         Row('ic_HW_PM','ic_ODM'),
                         Row('ic_Segment', 'ic_Chipset_Type'),
                         Row('ic_Phase'),
                         css_class='unsort no_title'
                         ),

                Fieldset('',
                         Row('ic_IC_Function','ic_IC_qty'),
                         Row('ic_Description'),
                         Row('ic_Defeature', 'ic_Full_Feature'),
                         css_class='unsort no_title'
                         ),

                Fieldset('',
                         Row('ic_Vendor_2nd', 'ic_Vendor_PN_2nd'),
                         Row('ic_Vendor_3rd', 'ic_Vendor_PN_3rd'),
                         Row('ic_Vendor_4th', 'ic_Vendor_PN_4th'),
                         css_class='unsort no_title'
                         ),
                ),
            Side(
                Fieldset('',
                         'ic_Vendor', 'ic_Vendor_PN', 'ic_Part_Owner', 'ic_HP_PN',
                         css_class='unsort no_title'
                         ),
            ),
            )
        return super(ICsAdmin, self).get_form_layout()

xadmin.site.register(ICs, ICsAdmin)